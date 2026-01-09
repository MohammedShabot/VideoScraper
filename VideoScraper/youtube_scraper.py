import time
import re
from typing import List, Dict, Optional
from urllib.parse import urlparse, parse_qs, unquote
import requests
import yt_dlp
from bs4 import BeautifulSoup
from urllib3.util.retry import Retry
from config import LoggerConfig
import random
from requests.adapters import HTTPAdapter
import pandas as pd
import itertools
import csv
import cleantext
from openpyxl import load_workbook, Workbook
import os

class VideoScraper:
    def __init__(self, rate_limit_delay: float = 2.0):
        self.rate_limit_delay = rate_limit_delay
        self.logger = LoggerConfig.setup_logger(__name__)


        self.stats = {
            "queries_processed": 0,
            "total_videos_found": 0,
            "youtube_videos": 0,
            "errors": 0,
        }

        self.csv_file = 'data/Scraping_Part1_keywords_extended.csv'
        
        self.stats = {
            "queries_processed": 0,
            "total_videos_found": 0,
            "youtube_videos": 0,
            "vimeo_videos": 0,
            "errors": 0,
        }

        self.queries = []
        self.scraped_urls = []
        self.fieldnames= ['query', 'platform', 'url', 'title', 'duration', 'view_count', 'description', 'uploader', 'upload_date']

        query_terms = pd.read_csv(self.csv_file)
        query_terms.fillna('', inplace=True)
        terms_emotion = set(query_terms.Emotion.unique())
        terms_subject = set(query_terms.Subject.unique())
        terms_setting = set(query_terms.Setting.unique())
        for s in [terms_emotion, terms_subject, terms_setting]:
            if '' in s:
                s.remove('')

        all_queries = list(itertools.product(terms_emotion, terms_setting, terms_subject))
        random.shuffle(all_queries)

        for p in all_queries:
            emo, setting, subj = p
            query = f"{emo} {subj} {setting}".strip()
            self.queries.append(query)

        self.logger.info(f"Initialized Youtube scraper with {len(self.queries)} queries")


    def _get_session(self):

        session = requests.Session()

        retry_strategy = Retry(
            total=3,
            backoff_factor=2,
            status_forcelist=[429, 500, 502, 503, 504], 
            allowed_methods=['GET', 'POST']
        )

        adapter = HTTPAdapter(max_retries=retry_strategy)
        session.mount("http://", adapter)
        session.mount("https://", adapter)

        return session
    
    def _get_random_headers(self):
        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Safari/605.1.15',
            'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        ]
        
        accept_languages = [
            'en-US,en;q=0.9',
            'en-GB,en;q=0.9',
            'en-US,en;q=0.9,nl;q=0.8',
            'en-US,en;q=0.8',
        ]
        
        return {
            'User-Agent': random.choice(user_agents),
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
            'Accept-Language': random.choice(accept_languages),
            'Accept-Encoding': 'gzip, deflate, br',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'none',
            'Cache-Control': 'max-age=0',
        }
    
    def make_csv_safe(self, text):
        return "".join(char for char in text if ord(char) <= 0xFFF)


    def scrape_youtube(self, query: str, max_results: int = 10) -> List[Dict]:
        """Scrape YouTube met yt-dlp"""
        self.logger.info(f"🔍 Starting YouTube search: '{query}'")
        self.logger.debug(f"   Max results: {max_results}")

        ydl_opts = {
            "quiet": True,
            "no_warnings": True,
            "extract_flat": False,  # attempt to fetch richer metadata in search results
            "skip_download": True,
            "noplaylist": True,
        }

        try:
            random_delay = random.uniform(1, 2)
            fixed_delay = 3

            print(f'Going to sleep for {fixed_delay + random_delay}')
            time.sleep(fixed_delay + random_delay)
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                search_url = f"ytsearch{max_results}:{query}"
                self.logger.debug(f"Search URL: {search_url}")

                results = ydl.extract_info(search_url, download=True)
                if not results:
                    self.logger.warning("yt-dlp returned None results")
                    return []

                entries = results.get("entries", []) or []
                self.logger.debug(f"Found {len(entries)} entries from YouTube")

                videos: List[Dict] = []
                for idx, entry in enumerate(entries):
                    if not entry:
                        continue

                
                    video_id = entry.get("id")
                    title = entry.get("title", "N/A")
                    duration = entry.get("duration", 0)
                    description = entry.get("description") or ""  # may be None depending on extractor

                    video_data = {
                        "platform": "youtube",
                        "url": f"https://www.youtube.com/watch?v={video_id}" if video_id else entry.get("webpage_url"),
                        "title": self.make_csv_safe(title),
                        "duration": duration or 0,
                        "view_count": entry.get("view_count"),
                        "description": self.make_csv_safe(description),
                        "uploader": self.make_csv_safe(entry.get("uploader")),
                        "upload_date": entry.get("upload_date"),
                    }

                    videos.append(video_data)
                    desc_len = len(description) if description else 0
                    self.logger.debug(f"   Video {idx + 1}: {title} ({duration}s) desc_len={desc_len}")

                self.logger.info(f" YouTube search complete: {len(videos)}")

                self.stats["youtube_videos"] += len(videos)
                return videos

        except Exception as e:
            self.logger.error(f"YouTube search failed: {e}", exc_info=True)
            self.stats["errors"] += 1
            return []
        
    def append_to_excel(self, file_path: str, rows: List[Dict]):
        if not rows:
            return

        # Zorg dat kolommen altijd in dezelfde volgorde komen
        def row_to_list(d: Dict):
            return [d.get(k) for k in self.fieldnames]

        if not os.path.exists(file_path):
            # Nieuw bestand + headers
            wb = Workbook()
            ws = wb.active
            ws.title = "videos"
            ws.append(self.fieldnames)  # header
            for r in rows:
                ws.append(row_to_list(r))
            wb.save(file_path)
            return

        # Bestaat al -> append
        wb = load_workbook(file_path)
        ws = wb.active  # of wb["videos"] als je zeker wil zijn van de sheetnaam
        for r in rows:
            ws.append(row_to_list(r))
        wb.save(file_path)

        
    def run_scraper(self):
        excel_path = "data/results/youtube_videos_scraped.xlsx"
        csv_path = "data/results/youtube_videos_scraped.csv"

        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=self.fieldnames)
            writer.writeheader()

            for query in self.queries:
                urls = self.scrape_youtube(query, 5)
                self.logger.info("Successfully scraped Youtube for some links, try to save now")

                if not urls:
                    continue

                enriched = []
                for video in urls:
                    video["query"] = query
                    enriched.append(video)

                    self.logger.info(f"Saving information for url: {video.get('url')}")
                    writer.writerow(video)

                self.append_to_excel(excel_path, enriched)
                self.logger.info("Video's saved in Excel and CSV format")


    
        

        